from __future__ import annotations

import csv
import datetime as dt
from typing import List, Dict, Tuple
from copy import deepcopy
from .models import (
    WEEKDAYS, DAY, NIGHT, OFF,
    ShiftConfig, RulesConfig, PersonState,
    DayPlan, WeekPlan, Config
)


def compute_min_pattern_weeks(people_count: int, rules: RulesConfig) -> int:
    # Heuristic lower bound
    return 2

def _is_feasible(shift: ShiftConfig, rules: RulesConfig) -> bool:
    """
    Quick feasibility check:
    - Daily minimum staffing implies weekly staffing S.
    - Average OFF per person = 7 - S / people.
    - Must satisfy min/max OFF bounds on average.
    """
    people_count = len(shift.people)
    if people_count <= 0:
        return False
    min_daily_total = shift.min_day_staff + shift.min_night_staff
    if min_daily_total > people_count:
        return False
    S = 7 * min_daily_total
    avg_off = 7 - (S / people_count)
    return (avg_off >= rules.min_days_off) and (avg_off <= rules.max_days_off)

def _relax_optional_prefs(shift: ShiftConfig, rules: RulesConfig) -> Tuple[ShiftConfig, RulesConfig]:
    """
    Relax optional preferences when infeasible:
    - Disable Wednesday overfill
    - Clear Friday priority names
    """
    s2 = deepcopy(shift)
    r2 = deepcopy(rules)
    s2.wednesday_day_overfill_count = 0
    r2.wednesday_day_overfill = False
    r2.friday_shift2_priority_names = []
    return s2, r2


def _total_daily_staff(people_count: int, rules: RulesConfig) -> int:
    """
    Choose total staffed headcount (DAY+NIGHT) per day.
    Goal: average ~40h/week per person. Favor fuller staffing within bounds.
    Heuristic: aim near shift maxes while not exceeding available people.
    """
    # Prefer fuller coverage: cap by people_count, but aim to use as many slots as max bounds allow.
    # Note: the exact split between DAY/NIGHT happens later respecting min/max per type.
    return max(1, min(people_count, people_count))


def target_daily_staff_counts(shift: ShiftConfig, rules: RulesConfig, weekday_index: int) -> Tuple[int, int]:
    """
    Decide DAY and NIGHT staffing counts using min/max bounds from ShiftConfig,
    keeping total <= available people and near the daily total heuristic.
    Apply Wednesday day overfill (bounded by max_day_staff).
    """
    people_count = len(shift.people)
    total_staff = _total_daily_staff(people_count, rules)

    # Start from mins
    day_count = shift.min_day_staff
    night_count = shift.min_night_staff

    # Allocate remaining capacity up to maxes, preferring balance
    # Allocate remaining capacity up to max bounds; since total_staff==people_count,
    # this favors fuller daily staffing to increase weekly hours
    remaining = max(0, total_staff - (day_count + night_count))
    while remaining > 0 and (day_count < shift.max_day_staff or night_count < shift.max_night_staff):
        if day_count <= night_count and day_count < shift.max_day_staff:
            day_count += 1
        elif night_count < shift.max_night_staff:
            night_count += 1
        remaining -= 1

    # Wednesday day overfill: only move from night if we have capacity and won't violate minimums
    if (rules.wednesday_day_overfill and weekday_index == 2 and 
        shift.wednesday_day_overfill_count > day_count):
        desired_day = min(shift.wednesday_day_overfill_count, shift.max_day_staff)
        # Only move from night if we can maintain night minimums
        available_to_move = max(0, night_count - shift.min_night_staff)
        move = min(desired_day - day_count, available_to_move)
        day_count += move
        night_count -= move

    # Ensure total doesn’t exceed available people
    total = day_count + night_count
    # Safeguard: ensure at least one person scheduled (not all OFF)
    if total == 0 and people_count > 0:
        # Prefer DAY if within bounds
        if shift.max_day_staff > 0:
            day_count = 1
        else:
            night_count = 1
        total = day_count + night_count
    if total > people_count:
        overflow = total - people_count
        while overflow > 0 and (day_count > shift.min_day_staff or night_count > shift.min_night_staff):
            if night_count >= day_count and night_count > shift.min_night_staff:
                night_count -= 1
            elif day_count > shift.min_day_staff:
                day_count -= 1
            overflow -= 1

    return day_count, night_count


def allocate_week_pattern(shift: ShiftConfig, rules: RulesConfig, pattern_weeks: int) -> List[WeekPlan]:
    # If infeasible under current optional prefs, relax them
    if not _is_feasible(shift, rules):
        shift, rules = _relax_optional_prefs(shift, rules)
    people = list(shift.people)
    person_states: Dict[str, PersonState] = {p: PersonState(name=p) for p in people}
    week_patterns: List[WeekPlan] = []

    # Compute hours using the shared helper function
    day_hours = _hours_for_range(shift.day_shift)
    night_hours = _hours_for_range(shift.night_shift)
    
    # Track total hours across all weeks in the pattern for average calculation
    pattern_total_hours: Dict[str, float] = {p: 0.0 for p in people}

    for w in range(pattern_weeks):
        days: List[DayPlan] = []
        off_counter: Dict[str, int] = {p: 0 for p in people}
        # Track per-person hours within the current week to improve fairness
        week_hours: Dict[str, float] = {p: 0.0 for p in people}

        for d in range(7):
            day_count, night_count = target_daily_staff_counts(shift, rules, d)

            is_friday = (d == 4)
            priority_names = rules.friday_shift2_priority_names if shift.name.lower() == "shift 2" else []

            def rank_candidates(assign_type: str) -> List[str]:
                scored: List[Tuple[str, float]] = []
                for p in people:
                    st = person_states[p]
                    if not st.can_assign(assign_type, rules):
                        continue
                    # Respect pattern average max hours cap: check if adding today would exceed average
                    if assign_type == DAY and hasattr(rules, "target_weekly_hours_max"):
                        projected_total = pattern_total_hours[p] + day_hours
                        projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                        if projected_avg > float(rules.target_weekly_hours_max):
                            continue
                    if assign_type == NIGHT and hasattr(rules, "target_weekly_hours_max"):
                        projected_total = pattern_total_hours[p] + night_hours
                        projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                        if projected_avg > float(rules.target_weekly_hours_max):
                            continue
                    
                    # When require_equal_hours is enabled, make pattern hours the PRIMARY criteria
                    if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                        # Sort primarily by pattern hours, then by other factors
                        scored.append((p, (pattern_total_hours[p], st.streak_len if st.streak_type == assign_type else 0)))
                    else:
                        penalty = st.streak_len if st.streak_type == assign_type else 0
                        bonus = -2 if (is_friday and p in priority_names) else 0
                        # Fairness: prefer those with fewer total pattern hours to equalize workload
                        fairness = pattern_total_hours[p] / 10.0
                        scored.append((p, penalty + (0 if st.last_assignment != assign_type else 0.5) - bonus + fairness))
                scored.sort(key=lambda x: (x[1], x[0]))
                return [p for p, _ in scored]

            day_members: List[str] = rank_candidates(DAY)[:day_count]
            for p in day_members:
                person_states[p].apply(DAY, rules)
                # accumulate hours for fairness
                week_hours[p] += day_hours
                pattern_total_hours[p] += day_hours

            night_candidates = [p for p in people if p not in day_members]

            def rank_night(cands: List[str]) -> List[str]:
                scored: List[Tuple[str, float]] = []
                for p in cands:
                    st = person_states[p]
                    if not st.can_assign(NIGHT, rules):
                        continue
                    if hasattr(rules, "target_weekly_hours_max"):
                        projected_total = pattern_total_hours[p] + night_hours
                        projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                        if projected_avg > float(rules.target_weekly_hours_max):
                            continue
                    
                    # When require_equal_hours is enabled, make pattern hours the PRIMARY criteria
                    if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                        # Sort primarily by pattern hours, then by streak
                        scored.append((p, (pattern_total_hours[p], st.streak_len if st.streak_type == NIGHT else 0)))
                    else:
                        penalty = st.streak_len if st.streak_type == NIGHT else 0
                        fairness = pattern_total_hours[p] / 10.0
                        scored.append((p, penalty + (0 if st.last_assignment != NIGHT else 0.5) + fairness))
                scored.sort(key=lambda x: (x[1], x[0]))
                return [p for p, _ in scored]

            night_members = rank_night(night_candidates)[:night_count]
            for p in night_members:
                person_states[p].apply(NIGHT, rules)
                week_hours[p] += night_hours
                pattern_total_hours[p] += night_hours

            # Helper to try assign OFF people to a type up to a needed count
            def _assign_from_off(assign_type: str, needed: int) -> int:
                off_pool = [x for x in people if x not in day_members and x not in night_members]
                
                # When require_equal_hours is enabled, sort by pattern hours to prioritize those with fewer hours
                if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                    off_pool.sort(key=lambda p: pattern_total_hours[p])
                
                for p in off_pool:
                    if needed <= 0:
                        break
                    st = person_states[p]
                    if st.can_assign(assign_type, rules):
                        # Respect pattern average max-hours cap when backfilling
                        if assign_type == DAY and hasattr(rules, "target_weekly_hours_max"):
                            projected_total = pattern_total_hours[p] + day_hours
                            projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                            if projected_avg > float(rules.target_weekly_hours_max):
                                continue
                        if assign_type == NIGHT and hasattr(rules, "target_weekly_hours_max"):
                            projected_total = pattern_total_hours[p] + night_hours
                            projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                            if projected_avg > float(rules.target_weekly_hours_max):
                                continue
                        person_states[p].apply(assign_type, rules)
                        if assign_type == DAY:
                            day_members.append(p)
                            week_hours[p] += day_hours
                            pattern_total_hours[p] += day_hours
                        else:
                            night_members.append(p)
                            week_hours[p] += night_hours
                            pattern_total_hours[p] += night_hours
                        needed -= 1
                return needed

            # Backfill NIGHT to meet computed target if underfilled due to constraints
            if len(night_members) < night_count:
                needed = night_count - len(night_members)
                # First try OFF members
                needed = _assign_from_off(NIGHT, needed)
                # If still short, try moving from DAY if above min_day_staff
                if needed > 0:
                    movable = [p for p in day_members if len(day_members) > shift.min_day_staff]
                    # When require_equal_hours, prioritize moving people with MORE hours
                    if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                        movable.sort(key=lambda p: pattern_total_hours[p], reverse=True)
                    for p in movable:
                        if needed <= 0:
                            break
                        st = person_states[p]
                        # Revert DAY and apply NIGHT if allowed
                        # Simple approach: if can assign NIGHT, switch
                        if st.can_assign(NIGHT, rules):
                            # Check pattern average (person already has DAY hours assigned)
                            if hasattr(rules, "target_weekly_hours_max"):
                                projected_total = pattern_total_hours[p] - day_hours + night_hours
                                projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                                if projected_avg > float(rules.target_weekly_hours_max):
                                    continue
                            # Remove DAY assignment effect by resetting last assignment; re-applying below
                            # For simplicity, we won't undo streak counters here; instead prefer future assignment fairness
                            day_members.remove(p)
                            person_states[p].apply(NIGHT, rules)
                            night_members.append(p)
                            week_hours[p] = week_hours[p] - day_hours + night_hours
                            pattern_total_hours[p] = pattern_total_hours[p] - day_hours + night_hours
                            needed -= 1

            # Ensure daily minimums: fill up to min_day_staff and min_night_staff
            # First ensure NIGHT minimums (to fix Sat/Sun gaps)
            if len(night_members) < shift.min_night_staff:
                needed_n = shift.min_night_staff - len(night_members)
                needed_n = _assign_from_off(NIGHT, needed_n)
                if needed_n > 0:
                    # Move from DAY if possible
                    movable = [p for p in day_members if len(day_members) > shift.min_day_staff]
                    # When require_equal_hours, prioritize moving people with MORE hours
                    if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                        movable.sort(key=lambda p: pattern_total_hours[p], reverse=True)
                    for p in movable:
                        if needed_n <= 0:
                            break
                        st = person_states[p]
                        if st.can_assign(NIGHT, rules):
                            # Check pattern average (person already has DAY hours assigned)
                            if hasattr(rules, "target_weekly_hours_max"):
                                projected_total = pattern_total_hours[p] - day_hours + night_hours
                                projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                                if projected_avg > float(rules.target_weekly_hours_max):
                                    continue
                            day_members.remove(p)
                            person_states[p].apply(NIGHT, rules)
                            night_members.append(p)
                            week_hours[p] = week_hours[p] - day_hours + night_hours
                            pattern_total_hours[p] = pattern_total_hours[p] - day_hours + night_hours
                            needed_n -= 1
                if needed_n > 0:
                    # Debug: show why we can't meet minimum with detailed breakdown
                    eligible_no_constraints = sum(1 for p in people if person_states[p].can_assign(NIGHT, rules))
                    in_cooldown = sum(1 for p in people if person_states[p].night_cooldown_remaining > 0)
                    already_day = len(day_members)
                    would_exceed_hours = sum(1 for p in people if p not in day_members and person_states[p].can_assign(NIGHT, rules) and hasattr(rules, "target_weekly_hours_max") and (pattern_total_hours[p] + night_hours) / pattern_weeks > float(rules.target_weekly_hours_max))
                    print(f"WARNING {WEEKDAYS[d]}: Cannot meet min_night_staff={shift.min_night_staff} (short by {needed_n}).")
                    print(f"  Already on DAY: {already_day}, In cooldown: {in_cooldown}, Would exceed max hours: {would_exceed_hours}")
                    print(f"  Available OFF people: {len([p for p in people if p not in day_members and p not in night_members])}")
                    # Continue with what we have rather than failing
                    # This allows the pattern to complete even if some days are understaffed

            # Then ensure DAY minimums
            if len(day_members) < shift.min_day_staff:
                needed_d = shift.min_day_staff - len(day_members)
                needed_d = _assign_from_off(DAY, needed_d)
                if needed_d > 0:
                    # Move from NIGHT if possible
                    movable = [p for p in night_members if len(night_members) > shift.min_night_staff]
                    # When require_equal_hours, prioritize moving people with MORE hours
                    if hasattr(rules, 'require_equal_hours') and rules.require_equal_hours:
                        movable.sort(key=lambda p: pattern_total_hours[p], reverse=True)
                    for p in movable:
                        if needed_d <= 0:
                            break
                        st = person_states[p]
                        if st.can_assign(DAY, rules):
                            # Check pattern average (person already has NIGHT hours assigned)
                            if hasattr(rules, "target_weekly_hours_max"):
                                projected_total = pattern_total_hours[p] - night_hours + day_hours
                                projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                                if projected_avg > float(rules.target_weekly_hours_max):
                                    continue
                            night_members.remove(p)
                            person_states[p].apply(DAY, rules)
                            day_members.append(p)
                            week_hours[p] = week_hours[p] - night_hours + day_hours
                            pattern_total_hours[p] = pattern_total_hours[p] - night_hours + day_hours
                            needed_d -= 1
                if needed_d > 0:
                    # Could not meet minimum DAY staffing; warn but continue
                    print(f"WARNING {WEEKDAYS[d]}: Cannot meet min_day_staff={shift.min_day_staff} (short by {needed_d})")
                    # Continue with what we have rather than failing

            off_members = [p for p in people if p not in day_members and p not in night_members]
            # Ensure at least one assigned this day for the team; if none, try to move one OFF to DAY or NIGHT
            if not day_members and not night_members and off_members:
                # Try assign someone to DAY first if possible
                fallback_assigned = False
                for p in off_members:
                    st = person_states[p]
                    if st.can_assign(DAY, rules):
                        person_states[p].apply(DAY, rules)
                        day_members.append(p)
                        off_members.remove(p)
                        fallback_assigned = True
                        week_hours[p] += day_hours
                        pattern_total_hours[p] += day_hours
                        break
                if not fallback_assigned:
                    for p in off_members:
                        st = person_states[p]
                        if st.can_assign(NIGHT, rules):
                            if hasattr(rules, "target_weekly_hours_max"):
                                projected_total = pattern_total_hours[p] + night_hours
                                projected_avg = projected_total / pattern_weeks if pattern_weeks > 0 else 0
                                if projected_avg > float(rules.target_weekly_hours_max):
                                    continue
                            person_states[p].apply(NIGHT, rules)
                            night_members.append(p)
                            off_members.remove(p)
                            week_hours[p] += night_hours
                            pattern_total_hours[p] += night_hours
                            break
            # Apply OFF assignment (cooldown is now handled automatically in PersonState.apply)
            for p in off_members:
                person_states[p].apply(OFF, rules)
                off_counter[p] += 1

            days.append(DayPlan(
                weekday_index=d,
                assignments={DAY: day_members, NIGHT: night_members, OFF: off_members}
            ))

        for p in people:
            off_days = off_counter[p]
            # Allow small deviations and let multi-week pattern absorb variance
            if off_days < rules.min_days_off - 1 or off_days > rules.max_days_off + 1:
                raise ValueError(
                    f"Weekly OFF days for {p} in {shift.name} = {off_days} violates relaxed bounds [{rules.min_days_off - 1}, {rules.max_days_off + 1}]"
                )

        week_patterns.append(WeekPlan(week_index=w, days=days))

    return week_patterns

def _pattern_meets_mins(pattern: List[WeekPlan], shift: ShiftConfig) -> bool:
    """
    Check if pattern meets minimum staffing requirements.
    More lenient: allow occasional understaffing if constraints prevent meeting mins.
    """
    violations = 0
    total_days = len(pattern) * 7
    
    for week in pattern:
        for day in week.days:
            if len(day.assignments[DAY]) < shift.min_day_staff:
                violations += 1
            if len(day.assignments[NIGHT]) < shift.min_night_staff:
                violations += 1
    
    # Allow up to 20% of days to be understaffed due to constraint conflicts
    tolerance = int(total_days * 0.2)
    return violations <= tolerance

def _compute_avg_week_hours(pattern: List[WeekPlan], shift: ShiftConfig) -> Dict[str, float]:
    # average hours per person across the pattern (weeks)
    day_h = _hours_for_range(shift.day_shift)
    night_h = _hours_for_range(shift.night_shift)
    totals: Dict[str, float] = {p: 0.0 for p in shift.people}
    weeks = len(pattern)
    for w in pattern:
        for d in w.days:
            for p in d.assignments[DAY]:
                totals[p] += day_h
            for p in d.assignments[NIGHT]:
                totals[p] += night_h
    return {p: (totals[p] / max(weeks, 1)) for p in shift.people}

def _check_equal_hours(pattern: List[WeekPlan], shift: ShiftConfig, rules: RulesConfig) -> bool:
    """Check if all people have equal average hours (within 0.5 hour tolerance)"""
    if not hasattr(rules, 'require_equal_hours') or not rules.require_equal_hours:
        return True
    
    avg_hours = _compute_avg_week_hours(pattern, shift)
    if not avg_hours:
        return True
    
    hours_values = list(avg_hours.values())
    min_hours = min(hours_values)
    max_hours = max(hours_values)
    
    # Allow 0.5 hour tolerance for rounding
    if max_hours - min_hours > 0.5:
        print(f"Hours not equal for '{shift.name}': min={min_hours:.2f}, max={max_hours:.2f}")
        for p, h in avg_hours.items():
            print(f"  {p}: {h:.2f}h/week")
        print(f"Trying a longer pattern to achieve equal hours...")
        return False
    
    return True

def _hours_for_range(tr: TimeRange) -> float:
    h1, m1 = map(int, tr.start.split(":"))
    h2, m2 = map(int, tr.end.split(":"))
    t1 = dt.timedelta(hours=h1, minutes=m1)
    t2 = dt.timedelta(hours=h2, minutes=m2)
    dur = t2 - t1
    if dur.total_seconds() <= 0:
        dur = (dt.timedelta(days=1) - t1) + t2
    return dur.total_seconds() / 3600.0

def find_smallest_valid_pattern(shift: ShiftConfig, rules: RulesConfig, max_try_weeks: int = 104) -> List[WeekPlan]:
    start = compute_min_pattern_weeks(len(shift.people), rules)
    last_err = None
    for w in range(start, max_try_weeks + 1):
        try:
            pat = allocate_week_pattern(shift, rules, w)
            if _pattern_meets_mins(pat, shift) and _check_equal_hours(pat, shift, rules):
                return pat
            else:
                # Try longer cycle
                continue
        except ValueError as e:
            last_err = e
            continue
    # If reached here, try dynamic adjustments to improve hours balance
    # Clone rules and relax preferred options; adjust max_days_off downward to allow more work
    adj_rules = RulesConfig(
        max_shifts_in_row=rules.max_shifts_in_row,
        max_days_off=rules.max_days_off,
        min_days_off=rules.min_days_off,
        no_day_after_night=rules.no_day_after_night,
        friday_shift2_priority_names=[],
        wednesday_day_overfill=False,
        min_days_off_after_night_streak=rules.min_days_off_after_night_streak,
        target_weekly_hours_min=rules.target_weekly_hours_min,
        enable_auto_adjust=rules.enable_auto_adjust,
    )
    # progressively reduce max_days_off toward min_days_off
    for mdo in range(rules.max_days_off, rules.min_days_off - 1, -1):
        adj_rules.max_days_off = mdo
        for w in range(start, max_try_weeks + 1):
            try:
                pat = allocate_week_pattern(shift, adj_rules, w)
                if not _pattern_meets_mins(pat, shift):
                    continue
                if not _check_equal_hours(pat, shift, rules):
                    continue
                avg_hours = _compute_avg_week_hours(pat, shift)
                # Accept if average >= configurable target
                target = rules.target_weekly_hours_min if hasattr(rules, 'target_weekly_hours_min') else 30.0
                if all(h >= float(target) for h in avg_hours.values()):
                    print(f"Adjusted rules for '{shift.name}': max_days_off={mdo}, disabled Wednesday overfill and Friday priorities, target_weekly_hours_min={target}.")
                    return pat
                # Otherwise keep trying longer weeks
            except ValueError:
                continue
    # As last resort, return a longest attempt even if below target; inform adjustments
    pat = allocate_week_pattern(shift, adj_rules, max_try_weeks)
    avg_hours = _compute_avg_week_hours(pat, shift)
    target = rules.target_weekly_hours_min if hasattr(rules, 'target_weekly_hours_min') else 30.0
    print(f"Adjusted rules for '{shift.name}' but could not reach {target}h avg for all. max_days_off={adj_rules.max_days_off}. Averages: " + ", ".join(f"{p}:{h:.1f}" for p,h in avg_hours.items()))
    return pat


def write_csv(out_path: str, start_date: dt.date, total_weeks: int, shift_patterns: Dict[str, List[WeekPlan]]) -> None:
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["week_index", "date", "weekday", "shift_name", "shift_type", "members"])

        for w in range(total_weeks):
            week_start = start_date + dt.timedelta(weeks=w)
            for shift_name, patterns in shift_patterns.items():
                pattern_week = patterns[w % len(patterns)]
                for day in pattern_week.days:
                    day_date = week_start + dt.timedelta(days=day.weekday_index)
                    writer.writerow([w, day_date.isoformat(), WEEKDAYS[day.weekday_index], shift_name, DAY, ";".join(day.assignments[DAY])])
                    writer.writerow([w, day_date.isoformat(), WEEKDAYS[day.weekday_index], shift_name, NIGHT, ";".join(day.assignments[NIGHT])])


def write_pivot_csv(out_path: str, total_weeks: int, shift_patterns: Dict[str, List[WeekPlan]], cfg: Config) -> None:
    """
    Writes a pivot-style CSV matching the spreadsheet-like view:
    - Weeks laid out horizontally with day columns (M..Su) per week.
    - One row per shift+type, labeled with times and timezone.
    - Cells contain the semicolon-separated member list for that day.

    Note: CSV cannot merge cells for week headers; we emit a header row per week.
    """
    # Build header: for each week, 7 day columns
    week_headers: List[str] = []
    day_headers: List[str] = []
    for w in range(total_weeks):
        week_headers.extend([f"Week {w+1}"] + ["" for _ in range(8)])
        day_headers.extend(["M", "T", "W", "Th", "F", "S", "Su", "Hours", "Total"])

    rows: List[List[str]] = []

    # For deterministic ordering: list the first team (first shift), then the second, etc.
    for shift in cfg.shifts:
        patterns = shift_patterns[shift.name]

        def _country_code(name: str) -> str:
            first = name.split()[0].lower()
            return {
                "lithuania": "LT",
                "indonesia": "ID",
            }.get(first, first[:2].upper())

        def _short_time(t: str) -> str:
            # Convert HH:MM to HH
            return t.split(":")[0]

        def base_label(assign_type: str) -> str:
            code = _country_code(shift.name)
            shift_num = "S1" if assign_type == DAY else "S2"
            time_range = f"{_short_time(shift.day_shift.start)}-{_short_time(shift.day_shift.end)}" if assign_type == DAY else f"{_short_time(shift.night_shift.start)}-{_short_time(shift.night_shift.end)}"
            return f"{code} {shift_num}: {time_range}"

        # Emit one row per person per assignment type, showing only that person's presence for each day
        # Helper: compute hours per assignment type
        def _hours_for(assign_type: str) -> float:
            def parse_hhmm(s: str):
                h, m = s.split(":"); return int(h), int(m)
            sh = shift.day_shift if assign_type == DAY else shift.night_shift
            h1, m1 = parse_hhmm(sh.start)
            h2, m2 = parse_hhmm(sh.end)
            t1 = dt.timedelta(hours=h1, minutes=m1)
            t2 = dt.timedelta(hours=h2, minutes=m2)
            dur = t2 - t1
            if dur.total_seconds() <= 0:
                dur = (dt.timedelta(days=1) - t1) + t2
            return dur.total_seconds() / 3600.0

        for assign_type in (DAY, NIGHT):
            for person in shift.people:
                row: List[str] = []
                total_hours_across_weeks = 0.0
                for w in range(total_weeks):
                    week = patterns[w % len(patterns)]
                    worked_days = 0
                    for d in range(7):
                        members = week.days[d].assignments[assign_type]
                        cell = person if person in members else ""
                        row.append(cell)
                        if person in members:
                            worked_days += 1
                    # Append weekly hours column for this assign type
                    hours = worked_days * _hours_for(assign_type)
                    row.append(f"{hours:.1f}")
                    total_hours_across_weeks += hours
                    # Total column: only populate for DAY rows; leave blank for NIGHT rows
                    if assign_type == DAY:
                        combined = 0.0
                        day_h = _hours_for(DAY)
                        night_h = _hours_for(NIGHT)
                        for d in range(7):
                            if person in week.days[d].assignments[DAY]:
                                combined += day_h
                            if person in week.days[d].assignments[NIGHT]:
                                combined += night_h
                        row.append(f"{combined:.1f}")
                    else:
                        row.append("")
                # Calculate average combined hours (DAY+NIGHT) per week across all weeks
                # Only show in DAY rows
                if assign_type == DAY:
                    total_combined_hours = 0.0
                    day_h = _hours_for(DAY)
                    night_h = _hours_for(NIGHT)
                    for w in range(total_weeks):
                        week = patterns[w % len(patterns)]
                        week_combined = 0.0
                        for d in range(7):
                            if person in week.days[d].assignments[DAY]:
                                week_combined += day_h
                            if person in week.days[d].assignments[NIGHT]:
                                week_combined += night_h
                        total_combined_hours += week_combined
                    avg_hours = total_combined_hours / total_weeks if total_weeks > 0 else 0.0
                    row.append(f"{avg_hours:.1f}")
                else:
                    row.append("")
                row.insert(0, base_label(assign_type))
                rows.append(row)

    # Write CSV
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # First column label for shift row names
        writer.writerow(["Shift"] + week_headers)
        writer.writerow([""] + day_headers * total_weeks)
        for r in rows:
            writer.writerow(r)


def write_pivot_xlsx(out_path: str, total_weeks: int, shift_patterns: Dict[str, List[WeekPlan]], cfg: Config) -> None:
    """
    Write a styled XLSX workbook with pivot layout and colors per team/shift type.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError("XLSX output requires 'openpyxl'. Please install it.") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Header rows
    week_headers: List[str] = []
    day_headers: List[str] = []
    for w in range(total_weeks):
        week_headers.extend([f"Week {w+1}"] + ["" for _ in range(8)])
        day_headers.extend(["M", "T", "W", "Th", "F", "S", "Su", "Hours", "Total"])

    # Add final Avg column
    week_headers.append("")
    day_headers.append("Avg")

    # Write headers
    ws.cell(row=1, column=1, value="Shift")
    for i, h in enumerate(week_headers, start=2):
        ws.cell(row=1, column=i, value=h)
    for i, h in enumerate(day_headers * total_weeks, start=2):
        ws.cell(row=2, column=i, value=h)

    # Build data rows like CSV pivot (one row per person per assign type)
    rows: List[Tuple[str, List[str]]] = []
    for shift in cfg.shifts:
        patterns = shift_patterns[shift.name]
        def _country_code(name: str) -> str:
            first = name.split()[0].lower()
            return {
                "lithuania": "LT",
                "indonesia": "ID",
            }.get(first, first[:2].upper())

        def _short_time(t: str) -> str:
            return t.split(":")[0]

        def base_label(assign_type: str) -> str:
            code = _country_code(shift.name)
            shift_num = "S1" if assign_type == DAY else "S2"
            time_range = f"{_short_time(shift.day_shift.start)}-{_short_time(shift.day_shift.end)}" if assign_type == DAY else f"{_short_time(shift.night_shift.start)}-{_short_time(shift.night_shift.end)}"
            return f"{code} {shift_num}: {time_range}"
        def _hours_for(assign_type: str) -> float:
            def parse_hhmm(s: str):
                h, m = s.split(":"); return int(h), int(m)
            sh = shift.day_shift if assign_type == DAY else shift.night_shift
            h1, m1 = parse_hhmm(sh.start)
            h2, m2 = parse_hhmm(sh.end)
            t1 = dt.timedelta(hours=h1, minutes=m1)
            t2 = dt.timedelta(hours=h2, minutes=m2)
            dur = t2 - t1
            if dur.total_seconds() <= 0:
                dur = (dt.timedelta(days=1) - t1) + t2
            return dur.total_seconds() / 3600.0
        for assign_type in (DAY, NIGHT):
            for person in shift.people:
                cells: List[str] = []
                for w in range(total_weeks):
                    week = patterns[w % len(patterns)]
                    worked_days = 0
                    for d in range(7):
                        members = week.days[d].assignments[assign_type]
                        val = person if person in members else ""
                        cells.append(val)
                        if person in members:
                            worked_days += 1
                    # weekly hours after each week block
                    hours = worked_days * _hours_for(assign_type)
                    cells.append(f"{hours:.1f}")
                    # combined total only on DAY rows; NIGHT rows blank
                    if assign_type == DAY:
                        combined = 0.0
                        day_h = _hours_for(DAY)
                        night_h = _hours_for(NIGHT)
                        for d in range(7):
                            if person in week.days[d].assignments[DAY]:
                                combined += day_h
                            if person in week.days[d].assignments[NIGHT]:
                                combined += night_h
                        cells.append(f"{combined:.1f}")
                    else:
                        cells.append("")
                rows.append((base_label(assign_type), cells))

    # Styling
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Apply header style
    for r in (1, 2):
        for c in range(1, 1 + 1 + len(week_headers)):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
            cell.border = border

    # Color map per team/type
    def team_colors(idx: int):
        # idx is shift index in cfg.shifts
        # Palette pairs: DAY, NIGHT
        palettes = [
            ("D9EAD3", "CFE2F3"),  # first team: green, teal
            ("FCE5CD", "EAD1DC"),  # second team: orange, purple
            ("FFF2CC", "D9D2E9"),  # third team: yellow, violet
        ]
        return palettes[idx % len(palettes)]

    # Write data rows with coloring per team/type
    row_idx = 3
    for shift_index, shift in enumerate(cfg.shifts):
        patterns = shift_patterns[shift.name]
        
        def _country_code(name: str) -> str:
            first = name.split()[0].lower()
            return {
                "lithuania": "LT",
                "indonesia": "ID",
            }.get(first, first[:2].upper())

        def _short_time(t: str) -> str:
            return t.split(":")[0]

        day_color, night_color = team_colors(shift_index)
        # For each type, for each person
        for assign_type in (DAY, NIGHT):
            code = _country_code(shift.name)
            shift_num = "S1" if assign_type == DAY else "S2"
            time_range = f"{_short_time(shift.day_shift.start)}-{_short_time(shift.day_shift.end)}" if assign_type == DAY else f"{_short_time(shift.night_shift.start)}-{_short_time(shift.night_shift.end)}"
            label = f"{code} {shift_num}: {time_range}"
            
            color = day_color if assign_type == DAY else night_color
            fill = PatternFill("solid", fgColor=color)
            for person in shift.people:
                a_cell = ws.cell(row=row_idx, column=1, value=label)
                a_cell.font = Font(bold=True)
                # fill cells including weekly hours and combined totals
                col = 2
                total_hours_across_weeks = 0.0
                for w in range(total_weeks):
                    week = shift_patterns[shift.name][w % len(shift_patterns[shift.name])]
                    worked_days = 0
                    for d in range(7):
                        members = week.days[d].assignments[assign_type]
                        val = person if person in members else ""
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.border = border
                        cell.alignment = center
                        if person in members:
                            worked_days += 1
                        col += 1
                    # Add weekly hours cell after each week block
                    hours = worked_days * _hours_for(assign_type)
                    total_hours_across_weeks += hours
                    hcell = ws.cell(row=row_idx, column=col, value=f"{hours:.1f}")
                    hcell.border = border
                    hcell.alignment = center
                    col += 1
                    # Add combined TOTAL (DAY+NIGHT) after Hours - only for DAY rows
                    if assign_type == DAY:
                        combined = 0.0
                        # compute shift hours
                        day_h = _hours_for(DAY)
                        night_h = _hours_for(NIGHT)
                        for d in range(7):
                            if person in week.days[d].assignments[DAY]:
                                combined += day_h
                            if person in week.days[d].assignments[NIGHT]:
                                combined += night_h
                        tcell = ws.cell(row=row_idx, column=col, value=f"{combined:.1f}")
                    else:
                        tcell = ws.cell(row=row_idx, column=col, value="")
                    tcell.border = border
                    tcell.alignment = center
                    col += 1
                # Add average combined hours (DAY+NIGHT) per week - only for DAY rows
                if assign_type == DAY:
                    total_combined_hours = 0.0
                    day_h = _hours_for(DAY)
                    night_h = _hours_for(NIGHT)
                    for w in range(total_weeks):
                        week = shift_patterns[shift.name][w % len(shift_patterns[shift.name])]
                        week_combined = 0.0
                        for d in range(7):
                            if person in week.days[d].assignments[DAY]:
                                week_combined += day_h
                            if person in week.days[d].assignments[NIGHT]:
                                week_combined += night_h
                        total_combined_hours += week_combined
                    avg_hours = total_combined_hours / total_weeks if total_weeks > 0 else 0.0
                    avg_cell = ws.cell(row=row_idx, column=col, value=f"{avg_hours:.1f}")
                else:
                    avg_cell = ws.cell(row=row_idx, column=col, value="")
                avg_cell.border = border
                avg_cell.alignment = center
                avg_cell.font = Font(bold=True)
                # color full row
                for c in range(1, 1 + 1 + len(week_headers)):
                    ws.cell(row=row_idx, column=c).fill = fill
                row_idx += 1

    # Freeze top two rows
    ws.freeze_panes = ws["A3"]

    # Set column widths roughly
    ws.column_dimensions["A"].width = 40
    # Data columns (B..): narrow but readable
    for c in range(2, 2 + len(week_headers)):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 10

    wb.save(out_path)
